import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes
import re
from config import TOKEN
import database
import logging

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

user_states = {}

database.create_table()

def normalizar_prioridad(prioridad):
    prioridad = prioridad.strip().lower()

    if prioridad in ["alta", "a"]:
        return "Alta"
    elif prioridad in ["media", "m"]:
        return "Media"
    elif prioridad in ["baja", "b"]:
        return "Baja"
    else:
        return None

def normalizar_nombre(nombre):
    return nombre.strip().title()

def clasificar(nombre):
    nombre = nombre.lower()

    categorias = {
        "Alimentación": ["taco", "comida", "pizza", "hamburguesa"],
        "Transporte": ["uber", "taxi", "bus", "metro"],
        "Entretenimiento": ["netflix", "spotify", "cine"],
        "Compras": ["ropa", "amazon", "zapatos"],
        "Salud": ["farmacia", "doctor", "medicina"]
    }

    for categoria, keywords in categorias.items():
        if any(x in nombre for x in keywords):
            return categoria

    return "Otros"

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    nombre_usuario = update.effective_user.first_name

    await update.message.reply_text(
        f"Hola, {nombre_usuario}\n\n"
        "Soy tu bot de gastos creado por Leonardo Correa\n\n"
        "Puedes hacer:\n\n"
        "➤ Guardar gasto:\n"
        "Comida, Alta, 120\n\n"
        "➤ Ver total:\n/total\n\n"
        "➤ Ver categorías:\n/categorias\n\n"
        "➤ Exportar Excel:\n/exportar\n\n"
        "➤ Actualizar gasto:\n/update_gasto\n\n"
        "➤ Borrar gasto:\n/borrar_gasto"
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):

    text = update.message.text
    user_id = update.effective_user.id

    # UPDATE GASTO
    if user_states.get(user_id) == "esperando_update":

        try:
            nombre, precio_real = [x.strip() for x in text.split(",")]

            nombre = normalizar_nombre(nombre)

            precio_real = float(precio_real)

        except:
            await update.message.reply_text(
                "Formato incorrecto.\n\n"
                "Usa:\n"
                "Nombre, Precio_real"
            )
            return

        updated = database.update_gasto(
            user_id,
            nombre,
            precio_real
        )

        if updated:

            await update.message.reply_text(
                f"Actualizado\n\n"
                f"{updated[0]}\n\n"
                f"Estimado: ${updated[1]}\n"
                f"Real: ${precio_real}"
            )

        else:
            await update.message.reply_text(
                "No se encontró el gasto."
            )

        user_states.pop(user_id)

        return
    
    # EXPORTAR CON / SIN GRAFICAS
    if user_states.get(user_id) == "esperando_tipo_exportacion":

        respuesta = text.strip().lower()

        if respuesta not in ["si", "sí", "no"]:

            await update.message.reply_text(
                "Responde solamente:\nSI\nNO"
            )
            return

        incluir_graficas = respuesta in ["si", "sí"]

        user_states.pop(user_id)

        await generar_excel(
            update,
            incluir_graficas
        )

        return
    
    # CASO GENERAL
    try:
        # BORRAR GASTO
        if user_states.get(user_id) == "esperando_borrado":

            nombre, razon = [x.strip() for x in text.split(",")]

            nombre = normalizar_nombre(nombre)

            if razon.lower() not in ["no necesario", "otro"]:
                raise ValueError()

            deleted = database.delete_gasto(user_id, nombre)

            if deleted:

                await update.message.reply_text(
                    f"Eliminado\n\n"
                    f"{deleted[0]} - ${deleted[1]}\n\n"
                    f"Razón: {razon}"
                )

            else:

                await update.message.reply_text(
                    "No se encontró ese gasto."
                )

            user_states.pop(user_id)

            return

        # GUARDAR GASTO
        nombre, prioridad, monto = [
            x.strip() for x in text.split(",")
        ]

        nombre = normalizar_nombre(nombre)

        prioridad = normalizar_prioridad(prioridad)

        if prioridad is None:
            raise ValueError()

        monto = float(monto)

        categoria = clasificar(nombre)

        database.insert_gasto(
            user_id,
            nombre,
            prioridad,
            monto,
            categoria
        )

        await update.message.reply_text(
            f"Guardado\n\n"
            f"{nombre}\n"
            f"{categoria}\n"
            f"${monto}"
        )

    except:

        await update.message.reply_text(
            "Formato incorrecto\n\n"

            "Guardar gasto:\n"
            "Nombre, Prioridad, Monto\n\n"

            "Ejemplo:\n"
            "Comida, Alta, 120\n\n"

            "Borrar gasto:\n"
            "Nombre, Razón\n\n"

            "Razones válidas:\n"
            "No necesario / Otro"
        )

async def total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    total = database.get_total(user_id)

    await update.message.reply_text(f"Tu total es:\n\n ${total}")


async def categorias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    data = database.get_by_category(user_id)

    if not data:
        await update.message.reply_text("No tienes gastos aún.")
        return

    msg = "Tus gastos por categoría:\n\n"

    for cat, total in data:
        msg += f"{cat}: ${total}\n"

    await update.message.reply_text(msg)

async def exportar(update: Update, context: ContextTypes.DEFAULT_TYPE):

    user_id = update.effective_user.id

    user_states[user_id] = "esperando_tipo_exportacion"

    await update.message.reply_text(
        "¿Quieres el Excel con gráficas?\n\n"
        "Responde:\n"
        "SI\n"
        "NO"
    )

async def generar_excel(update, incluir_graficas=False):

    user_id = update.effective_user.id

    data = database.get_all_by_user(user_id)

    user = update.effective_user

    nombre_usuario = re.sub(
        r"[^a-zA-Z0-9_]",
        "",
        user.first_name
    )

    if not data:

        await update.message.reply_text(
            "No tienes gastos para exportar."
        )

        return

    # DATAFRAME
    df = pd.DataFrame(
        data,
        columns=[
            "Nombre",
            "Prioridad",
            "Monto",
            "Categoría",
            "Fecha",
            "Cumplido",
            "Precio Real"
        ]
    )

    df["Cumplido"] = df["Cumplido"].map({
        True: "Sí",
        False: "No"
    })

    df["Precio Real"] = df["Precio Real"].fillna(0)

    df["Diferencia"] = (
        df["Precio Real"] - df["Monto"]
    )

    # EXCEL EN MEMORIA
    file_stream = BytesIO()

    df.to_excel(file_stream, index=False)

    file_stream.seek(0)

    wb = load_workbook(file_stream)

    ws = wb.active

    # ESTILOS
    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    alta_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    media_fill = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )

    baja_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    # HEADER
    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

        cell.alignment = Alignment(
            horizontal="center"
        )

    # AUTO WIDTH
    for col in ws.columns:

        max_length = 0

        col_letter = col[0].column_letter

        for cell in col:

            if cell.value:

                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        ws.column_dimensions[col_letter].width = (
            max_length + 3
        )

    # COLORES PRIORIDAD
    for row in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row
    ):

        prioridad = str(row[1].value).lower()

        fill = None

        if prioridad == "alta":
            fill = alta_fill

        elif prioridad == "media":
            fill = media_fill

        elif prioridad == "baja":
            fill = baja_fill

        if fill:

            for cell in row:
                cell.fill = fill

    # RESUMEN
    last_row = ws.max_row + 2

    ws[f"A{last_row}"] = "RESUMEN"

    ws[f"A{last_row}"].font = Font(
        bold=True
    )

    ws[f"A{last_row+1}"] = "Total estimado"
    ws[f"B{last_row+1}"] = df["Monto"].sum()

    ws[f"A{last_row+2}"] = "Total real"
    ws[f"B{last_row+2}"] = df["Precio Real"].sum()

    ws[f"A{last_row+3}"] = "Diferencia total"
    ws[f"B{last_row+3}"] = df["Diferencia"].sum()

    # GRAFICAS
    if incluir_graficas:

        try:

            resumen_categoria = (
                df.groupby("Categoría")["Monto"]
                .sum()
                .reset_index()
            )

            # TABLA AUXILIAR
            ws["J1"] = "Categoría"
            ws["K1"] = "Monto"

            for idx, row_data in resumen_categoria.iterrows():

                ws[f"J{idx+2}"] = str(row_data["Categoría"])

                ws[f"K{idx+2}"] = float(row_data["Monto"])

            # REFERENCIAS
            labels = Reference(
                ws,
                min_col=10,
                min_row=2,
                max_row=len(resumen_categoria) + 1
            )

            data_ref = Reference(
                ws,
                min_col=11,
                min_row=1,
                max_row=len(resumen_categoria) + 1
            )

            # PIE CHART
            pie = PieChart()

            pie.add_data(
                data_ref,
                titles_from_data=True
            )

            pie.set_categories(labels)

            pie.title = "Gastos por categoría"

            pie.height = 10
            pie.width = 12

            ws.add_chart(pie, "M2")

            # BAR CHART

            bar = BarChart()

            bar.add_data(
                data_ref,
                titles_from_data=True
            )

            bar.set_categories(labels)

            bar.title = "Comparación de gastos"

            bar.y_axis.title = "Monto"

            bar.x_axis.title = "Categoría"

            bar.height = 10
            bar.width = 14

            ws.add_chart(bar, "M20")

        except Exception as e:

            print("ERROR GRAFICAS:", e)

    # GUARDAR
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    filename = f"gastos_{nombre_usuario}.xlsx"

    await update.message.reply_document(
        document=output,
        filename=filename
    )

async def borrar_gasto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    user_states[user_id] = "esperando_borrado"

    await update.message.reply_text(
        "Borrar gasto\n\n"
        "Formato:\n"
        "Nombre, Razón\n\n"
        "Ejemplo:\n"
        "Uber, No necesario"
    )

async def update_gasto_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    data = database.get_pendientes(user_id)

    if not data:
        await update.message.reply_text("No tienes gastos pendientes")
        return

    mensaje = "Gastos pendientes:\n\n"

    for nombre, monto in data:
        mensaje += f"{nombre} - ${monto}\n"

    mensaje += "\nFormato:\nNombre, Precio_real\nEjemplo:\nUber, 95"

    user_states[user_id] = "esperando_update"

    await update.message.reply_text(mensaje)

def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("total", total))
    app.add_handler(CommandHandler("categorias", categorias))
    app.add_handler(CommandHandler("exportar", exportar))
    app.add_handler(CommandHandler("borrar_gasto", borrar_gasto))
    app.add_handler(CommandHandler("update_gasto", update_gasto_cmd))
    app.add_handler(MessageHandler(filters.TEXT, handle_message))
    print("Bot multiusuario corriendo...")
    app.run_polling()

if __name__ == "__main__":
    main()